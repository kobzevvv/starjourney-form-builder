import openpyxl
import requests
import json

# Открываем Excel-файл
wb = openpyxl.load_workbook('Candidate journey Generator.xlsx')
ws = wb.active

# Получаем значения из C4 и D4
job_description = ws['C4'].value
must_haves = ws['D4'].value

# Пример зарплаты для теста
salary = 5100

# Формируем структуру, похожую на Typeform webhook
SALARY_FIELD_ID = "087f6491-3724-4cc0-8719-4415ae11e8e9"

data = {
    "form_response": {
        "answers": [
            {
                "field": {"id": SALARY_FIELD_ID, "type": "number"},
                "type": "number",
                "number": salary
            },
            {
                "field": {"id": "job_description", "type": "text"},
                "type": "text",
                "text": job_description
            },
            {
                "field": {"id": "must_haves", "type": "text"},
                "type": "text",
                "text": must_haves
            }
        ]
    }
}

# Отправляем POST-запрос к локальной функции
response = requests.post(
    "http://localhost:8080/salary-check",
    json=data
)

print("Status:", response.status_code)
print("Response:", response.json()) 